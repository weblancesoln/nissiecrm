"""
Services for lead import/export functionality.
"""
import csv
import io

try:
    from openpyxl import load_workbook, Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .models import Lead


def _normalize_col(s):
    return str(s).strip().lower().replace(' ', '_') if s else ''


def import_leads_from_file(file, user=None):
    from django.contrib.auth.models import User
    """
    Import leads from CSV or Excel file.
    Expected columns: first_name, last_name (or prospect_name/name for legacy),
                     phone_number, email, point_of_contact, etc.
    Returns (success_count, error_messages).
    """
    errors = []
    success_count = 0

    try:
        filename = file.name.lower()
        rows = []

        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
            if not rows:
                return 0, ['File is empty.']
            headers = [_normalize_col(h) for h in rows[0]]
            data_rows = rows[1:]
        elif filename.endswith(('.xlsx', '.xls')):
            if not HAS_OPENPYXL:
                return 0, ['Excel support requires openpyxl. Run: pip install openpyxl']
            wb = load_workbook(filename=file, read_only=True)
            ws = wb.active
            rows_list = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows_list:
                return 0, ['File is empty.']
            headers = [_normalize_col(h) for h in rows_list[0]]
            data_rows = [list(r) for r in rows_list[1:]]
        else:
            return 0, ['Unsupported file format. Use CSV or Excel.']

        column_map = {
            'first_name': ['first_name', 'firstname', 'given_name'],
            'last_name': ['last_name', 'lastname', 'surname', 'family_name'],
            'prospect_name': ['prospect_name', 'name', 'full_name', 'contact_name', 'customer_name'],
            'phone_number': ['phone_number', 'phone', 'tel', 'mobile', 'contact'],
            'email': ['email', 'e-mail', 'mail'],
            'point_of_contact': ['point_of_contact', 'contact_point', 'poc', 'referral'],
            'prospect_response': ['prospect_response', 'response', 'feedback'],
            'remarks': ['remarks', 'notes', 'comments'],
            'status': ['status'],
            'source': ['source', 'lead_source'],
            'color_code': ['color_code', 'color'],
            'assigned_to': ['assigned_to', 'assigned_staff', 'staff', 'assigned'],
        }

        def find_col(possible_names):
            for name in possible_names:
                if name in headers:
                    return headers.index(name)
            return None

        col_first = find_col(column_map['first_name'])
        col_last = find_col(column_map['last_name'])
        col_prospect = find_col(column_map['prospect_name'])
        if col_first is None and col_prospect is None:
            return 0, ['Required: "first_name" or "prospect_name"/"name" not found.']

        col_phone = find_col(column_map['phone_number'])
        col_email = find_col(column_map['email'])
        col_poc = find_col(column_map['point_of_contact'])
        col_response = find_col(column_map['prospect_response'])
        col_remarks = find_col(column_map['remarks'])
        col_status = find_col(column_map['status'])
        col_source = find_col(column_map['source'])
        col_color = find_col(column_map['color_code'])
        col_assigned = find_col(column_map['assigned_to'])

        def get_val(row, idx):
            if idx is not None and idx < len(row):
                v = row[idx]
                return str(v).strip() if v is not None else ''
            return ''

        valid_statuses = [s[0] for s in Lead.STATUS_CHOICES]
        valid_colors = [c[0] for c in Lead.COLOR_CHOICES if c[0]]

        for i, row in enumerate(data_rows):
            try:
                if col_first is not None:
                    first_name = get_val(row, col_first)
                    last_name = get_val(row, col_last) if col_last is not None else ''
                else:
                    full = get_val(row, col_prospect)
                    parts = full.split(None, 1)
                    first_name = parts[0][:100] if parts else ''
                    last_name = (parts[1][:100] if len(parts) > 1 else '')
                if not first_name:
                    errors.append(f"Row {i + 2}: Missing first name, skipped")
                    continue

                status_val = get_val(row, col_status)
                status = status_val.lower() if status_val.lower() in valid_statuses else 'new'

                color_val = get_val(row, col_color)
                color_code = color_val if color_val in valid_colors else ''

                assigned_to = None
                if col_assigned is not None:
                    staff_val = get_val(row, col_assigned)
                    if staff_val:
                        try:
                            assigned_to = User.objects.get(username__iexact=staff_val)
                        except User.DoesNotExist:
                            pass

                lead = Lead(
                    first_name=first_name[:100],
                    last_name=last_name[:100],
                    phone_number=get_val(row, col_phone)[:50],
                    email=get_val(row, col_email)[:254],
                    point_of_contact=get_val(row, col_poc)[:200],
                    prospect_response=get_val(row, col_response),
                    remarks=get_val(row, col_remarks),
                    source=get_val(row, col_source)[:100],
                    status=status,
                    color_code=color_code,
                    assigned_to=assigned_to,
                    created_by=user,
                )
                lead.save()
                success_count += 1
            except Exception as e:
                errors.append(f"Row {i + 2}: {str(e)}")

    except Exception as e:
        errors.append(f"File processing error: {str(e)}")

    return success_count, errors


def export_leads_to_csv(queryset):
    """Export leads to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'First Name', 'Last Name', 'Phone Number', 'Email', 'Point of Contact',
        'Prospect Response', 'Remarks', 'Status', 'Color Code', 'Source',
        'Assigned To', 'Created At', 'Updated At'
    ])
    for lead in queryset:
        writer.writerow([
            lead.first_name, lead.last_name, lead.phone_number, lead.email,
            lead.point_of_contact, lead.prospect_response, lead.remarks,
            lead.status, lead.color_code, lead.source,
            lead.assigned_to.username if lead.assigned_to else '',
            lead.created_at.strftime('%Y-%m-%d %H:%M'),
            lead.updated_at.strftime('%Y-%m-%d %H:%M'),
        ])
    return output.getvalue()


def export_leads_to_excel(queryset):
    """Export leads to Excel format. Requires openpyxl."""
    if not HAS_OPENPYXL:
        raise ImportError('Excel export requires openpyxl. Run: pip install openpyxl')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leads'
    headers = [
        'First Name', 'Last Name', 'Phone Number', 'Email', 'Point of Contact',
        'Prospect Response', 'Remarks', 'Status', 'Color Code', 'Source',
        'Assigned To', 'Created At', 'Updated At'
    ]
    ws.append(headers)
    for lead in queryset:
        ws.append([
            lead.first_name, lead.last_name, lead.phone_number, lead.email,
            lead.point_of_contact, lead.prospect_response, lead.remarks,
            lead.get_status_display(), lead.get_color_code_display() or '',
            lead.source,
            lead.assigned_to.username if lead.assigned_to else '',
            lead.created_at.strftime('%Y-%m-%d %H:%M'),
            lead.updated_at.strftime('%Y-%m-%d %H:%M'),
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
